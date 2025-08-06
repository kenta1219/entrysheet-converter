import json
import logging
import tempfile
import os
import re
from datetime import datetime
from typing import List, Dict, Optional
import pandas as pd
import openpyxl

logger = logging.getLogger(__name__)


class TemplateBasedFileProcessingRepository:
    """テンプレートベースのファイル処理リポジトリ"""
    
    def process_with_template_mapping(self, xlsb_file, template_file, template):
        """テンプレート固有のマッピングを使用してファイル処理を実行"""
        import tempfile
        import pandas as pd
        import openpyxl
        from openpyxl.utils import column_index_from_string
        
        # 1. xlsbファイルからデータを抽出
        extracted_data = self._extract_data_with_template_mapping(xlsb_file, template)
        
        # 2. テンプレートファイルにデータを書き込み
        result_content = self._write_data_with_template_mapping(template_file, extracted_data, template)
        
        return result_content
    
    def _extract_data_with_template_mapping(self, xlsb_file, template):
        """テンプレートマッピングに基づいてデータを抽出"""
        with tempfile.NamedTemporaryFile(suffix='.xlsb', delete=False) as temp_file:
            temp_file.write(xlsb_file.content)
            temp_file.flush()
            
            try:
                # xlsbファイルを読み込み
                df = pd.read_excel(
                    temp_file.name,
                    sheet_name=template.mapping.source_sheet,
                    engine='pyxlsb'
                )
                
                # テンプレートのマッピングに基づいてデータを抽出
                extracted_values = {}
                
                for cell_mapping in template.mapping.cell_mappings:
                    target_cell = cell_mapping.target
                    
                    if cell_mapping.type == "single":
                        # 単一セルの値を取得
                        value = self._get_cell_value(df, cell_mapping.source)
                        extracted_values[target_cell] = value
                        
                    elif cell_mapping.type == "concat_cells":
                        # 複数セルを連結
                        source_cells = cell_mapping.source.split('+')
                        values = []
                        
                        for source_cell in source_cells:
                            cell_value = self._get_cell_value(df, source_cell.strip())
                            
                            # フォーマットルールを適用
                            if hasattr(cell_mapping, 'format_rules') and cell_mapping.format_rules:
                                format_rule = cell_mapping.format_rules.get(source_cell.strip())
                                if format_rule:
                                    cell_value = self._apply_format_rule(cell_value, format_rule)
                            
                            if cell_value:  # 空でない値のみ追加
                                values.append(str(cell_value))
                        
                        # セパレーターで連結
                        separator = getattr(cell_mapping, 'separator', '')
                        extracted_values[target_cell] = separator.join(values)
                
                return extracted_values
                
            except Exception as e:
                raise Exception(f"データ抽出エラー: {str(e)}")
    
    def _get_cell_value(self, df, cell_ref):
        """DataFrameから指定セルの値を取得（ExtractionConfig準拠）"""
        try:
            # セル参照をパース（例：F40 -> 列F、行40）
            col_letter = ''.join(filter(str.isalpha, cell_ref))
            row_num = int(''.join(filter(str.isdigit, cell_ref)))
            
            # 列文字を数値インデックスに変換（F=6）
            from openpyxl.utils import column_index_from_string
            col_index = column_index_from_string(col_letter) - 1
            
            # DataFrameから値を取得（行は0ベースなので-1）
            if row_num - 1 < len(df) and col_index < len(df.columns):
                value = df.iloc[row_num - 1, col_index]
                
                # NaN、None、空文字列の処理
                if pd.isna(value) or value is None:
                    return ""
                
                # 数値の場合は適切に変換
                if isinstance(value, (int, float)):
                    if pd.isna(value):
                        return ""
                    # 整数として表現できる場合は整数に変換
                    if isinstance(value, float) and value.is_integer():
                        return str(int(value))
                    return str(value)
                
                # 文字列の場合はそのまま返す
                return str(value).strip()
            
            return ""
        except Exception as e:
            logger.debug(f"セル値取得エラー {cell_ref}: {str(e)}")
            return ""
    
    def _apply_format_rule(self, value, format_rule):
        """フォーマットルールを適用（ExtractionConfig準拠）"""
        if not value or not format_rule:
            return value
            
        try:
            if format_rule == "zenkaku_int":
                # 半角数字を全角数字に変換
                import unicodedata
                # まず半角数字のみ抽出
                digits_only = ''.join(filter(str.isdigit, str(value)))
                if digits_only:
                    # 半角数字を全角数字に変換
                    zenkaku_digits = ""
                    for digit in digits_only:
                        zenkaku_digits += chr(ord(digit) - ord('0') + ord('０'))
                    return zenkaku_digits
                return ""
                
            elif format_rule == "hankaku_int":
                # 半角数字のみ抽出して整数処理
                digits_only = ''.join(filter(str.isdigit, str(value)))
                if digits_only:
                    return str(int(digits_only))
                return ""
                
            return str(value)
            
        except Exception as e:
            logger.debug(f"フォーマットルール適用エラー {format_rule}: {str(e)}")
            return str(value)
    
    def _write_data_with_template_mapping(self, template_file, extracted_data, template):
        """テンプレートマッピングに基づいてデータを書き込み"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx') as temp_file:
            temp_file.write(template_file.content)
            temp_file.flush()
            
            try:
                workbook = openpyxl.load_workbook(temp_file.name)
                worksheet = workbook[template.mapping.target_sheet]
                
                # マッピングに基づいてデータを書き込み
                for target_cell, value in extracted_data.items():
                    try:
                        # セル参照をパース（例：E16 -> 列E、行16）
                        col_letter = ''.join(filter(str.isalpha, target_cell))
                        row_num = int(''.join(filter(str.isdigit, target_cell)))
                        
                        cell = worksheet[target_cell]
                        
                        # 結合セルの処理
                        merged_range = self._find_merged_cell_range(worksheet, cell)
                        if merged_range:
                            worksheet.unmerge_cells(str(merged_range))
                            cell.value = self._convert_value_for_cell(value)
                            worksheet.merge_cells(str(merged_range))
                        else:
                            cell.value = self._convert_value_for_cell(value)
                            
                    except Exception as e:
                        logger.warning(f"セル {target_cell} への書き込みをスキップ: {str(e)}")
                        continue
                
                # ワークブックを保存してバイト配列として返す
                with tempfile.NamedTemporaryFile() as output_temp:
                    workbook.save(output_temp.name)
                    workbook.close()
                    
                    output_temp.seek(0)
                    return output_temp.read()
                    
            except Exception as e:
                raise Exception(f"テンプレート書き込みエラー: {str(e)}")
    
    def _find_merged_cell_range(self, worksheet, cell):
        """指定されたセルが結合セルの一部かどうかを確認し、結合範囲を返す"""
        try:
            for merged_range in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return merged_range
            return None
        except Exception:
            return None
    
    def _convert_value_for_cell(self, value):
        """セル用の値に変換"""
        if not value or value == "0":
            return value if value else ""
        
        try:
            return float(value)
        except (ValueError, TypeError):
            return value if value else ""
    
    def extract_data_from_xlsb(self, xlsb_file):
        """xlsbファイルからデータを抽出（複数ファイル処理用）"""
        from ..domain.entities import ExtractedData, ExtractionConfig
        
        with tempfile.NamedTemporaryFile(suffix='.xlsb', delete=False) as temp_file:
            temp_file.write(xlsb_file.content)
            temp_file.flush()
            
            try:
                # xlsbファイルを読み込み
                df = pd.read_excel(
                    temp_file.name,
                    sheet_name="加盟店申込書_施設名",  # デフォルトシート名
                    engine='pyxlsb'
                )
                
                # ExtractionConfigを使用してデータを抽出
                extraction_config = ExtractionConfig()
                extracted_values = []
                
                for cell_ref in extraction_config.cell_references:
                    if cell_ref.is_sum:
                        # 合計計算
                        total = 0
                        for cell in cell_ref.cells:
                            value = self._get_cell_value(df, cell)
                            try:
                                total += float(value) if value else 0
                            except (ValueError, TypeError):
                                pass
                        extracted_values.append(str(total))
                        
                    elif cell_ref.is_concat:
                        # 文字列連結
                        values = []
                        for cell in cell_ref.cells:
                            value = self._get_cell_value(df, cell)
                            if value:
                                values.append(str(value))
                        extracted_values.append(cell_ref.separator.join(values))
                        
                    else:
                        # 単一セル
                        value = self._get_cell_value(df, cell_ref.cells[0])
                        extracted_values.append(str(value) if value else "")
                
                return ExtractedData(
                    values=extracted_values,
                    source_sheet="加盟店申込書_施設名",
                    source_references=extraction_config.cell_references
                )
                
            except Exception as e:
                raise Exception(f"xlsbデータ抽出エラー: {str(e)}")
            finally:
                try:
                    os.unlink(temp_file.name)
                except:
                    pass
    
    def write_multiple_rows_to_template(self, template_file, template_info, row_data_list):
        """テンプレートに複数行のデータを書き込み"""
        logger.info(f"複数行書き込み開始 - 対象行数: {len(row_data_list)}")
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_template:
            temp_template.write(template_file.content)
            temp_template_path = temp_template.name
        
        try:
            # Excelファイルを開く
            workbook = openpyxl.load_workbook(temp_template_path)
            worksheet = workbook[template_info.mapping.target_sheet]
            
            # 各行データを順次書き込み
            for row_data in row_data_list:
                self._write_single_row_data(worksheet, template_info.mapping, row_data)
            
            # 処理済みファイルを保存
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_output:
                workbook.save(temp_output.name)
                temp_output_path = temp_output.name
            
            # バイナリデータとして読み込み
            with open(temp_output_path, 'rb') as f:
                result_content = f.read()
            
            logger.info(f"複数行書き込み完了 - 書き込み行数: {len(row_data_list)}")
            return result_content
            
        except Exception as e:
            logger.error(f"複数行書き込みエラー: {str(e)}")
            raise
        finally:
            # 一時ファイルをクリーンアップ
            try:
                os.unlink(temp_template_path)
                if 'temp_output_path' in locals():
                    os.unlink(temp_output_path)
            except:
                pass
    
    def _write_single_row_data(self, worksheet, mapping, row_data):
        """単一行のデータを書き込み"""
        logger.debug(f"行データ書き込み - 行番号: {row_data.row_number}, 施設名: {row_data.facility_name}")
        
        # セルマッピングに従って各セルに値を書き込み
        for i, cell_mapping in enumerate(mapping.cell_mappings):
            if i < len(row_data.extracted_values):
                # 出力先セルの行番号を動的に調整
                target_cell = self._adjust_cell_row_number(
                    cell_mapping.target,
                    row_data.row_number
                )
                
                # セルに値を書き込み
                try:
                    cell = worksheet[target_cell]
                    value = row_data.extracted_values[i]
                    
                    # 結合セルの処理
                    merged_range = self._find_merged_cell_range(worksheet, cell)
                    if merged_range:
                        worksheet.unmerge_cells(str(merged_range))
                        cell.value = self._convert_value_for_cell(value)
                        worksheet.merge_cells(str(merged_range))
                    else:
                        cell.value = self._convert_value_for_cell(value)
                    
                    logger.debug(f"セル書き込み: {target_cell} = {value}")
                    
                except Exception as e:
                    logger.warning(f"セル {target_cell} への書き込みをスキップ: {str(e)}")
                    continue
    
    def _adjust_cell_row_number(self, original_cell: str, new_row: int) -> str:
        """セル参照の行番号を調整"""
        # セル参照から列文字と行番号を分離
        match = re.match(r'([A-Z]+)(\d+)', original_cell)
        if match:
            column = match.group(1)
            return f"{column}{new_row}"
        else:
            raise ValueError(f"無効なセル参照: {original_cell}")
    
    def validate_template_capacity(self, template_info, required_rows: int, start_row: int = 14):
        """テンプレートの容量チェック"""
        from ..domain.entities import ValidationResult
        
        max_excel_rows = 1048576  # Excelの最大行数
        end_row = start_row + required_rows - 1
        
        if end_row > max_excel_rows:
            return ValidationResult.invalid(
                f"必要な行数が多すぎます。最大{max_excel_rows - start_row + 1}行まで処理可能です。"
            )
        
        # テンプレート固有の制限チェック
        if required_rows > 1000:  # 例：1000行制限
            return ValidationResult.invalid(
                f"一度に処理できる最大行数は1000行です。現在の要求: {required_rows}行"
            )
        
        return ValidationResult.valid()


class StructuredLoggerRepository:
    """構造化ログリポジトリ"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def log_info(self, message: str, **kwargs):
        """情報ログを出力"""
        self.logger.info(message, extra=kwargs)
    
    def log_warning(self, message: str, **kwargs):
        """警告ログを出力"""
        self.logger.warning(message, extra=kwargs)
    
    def log_error(self, message: str, **kwargs):
        """エラーログを出力"""
        self.logger.error(message, extra=kwargs)
    
    def log_debug(self, message: str, **kwargs):
        """デバッグログを出力"""
        self.logger.debug(message, extra=kwargs)